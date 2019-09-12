import cv2
import numpy as np
import pyzbar.pyzbar as pyzbar
from openpyxl import load_workbook

cap = cv2.VideoCapture(0)
font = cv2.FONT_HERSHEY_PLAIN

while True:
    _, frame = cap.read()

    decodedObjects = pyzbar.decode(frame)

    cv2.imshow("Frame", frame)

    for obj in decodedObjects:
        h = int(obj.data) + 1
        wb = load_workbook(filename = 'excel.xlsx')
        sheet=wb['Planilha1']
        wb.active
        print(sheet.cell(row=h,column=2).value)
        img = (sheet.cell(row=h, column=3).value)
        img = cv2.imread(img, 1)
        cv2.imshow('', img)
        cv2.waitKey(0)
        cv2.destroyAllWindows()
        break
   
    key = cv2.waitKey(1)
    if key == 27:
        break

    elif key == ord('s'):
        cv2.imwrite('qrcode_result.jpg', frame)    
        break